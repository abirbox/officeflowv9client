"""CSV + PDF builders for Dispatch reports (permission-aware)."""
import csv
import io
import os
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image,
)

APP_TZ = ZoneInfo("Asia/Dhaka")


def _dhaka_now_str() -> str:
    """Current wall-clock time in Asia/Dhaka (UTC+6) for report footers."""
    return datetime.now(timezone.utc).astimezone(APP_TZ).strftime("%Y-%m-%d %H:%M")



def build_advance_statement_pdf(
    *,
    officer_name: str,
    client_name: str | None,
    entries: list,
    total_advanced: float,
    total_repaid: float,
    remaining_balance: float,
) -> bytes:
    """Per-officer advance salary statement: every advance & repayment with a
    running balance, plus totals. ``entries`` items: {entry_date, type,
    note, amount, balance_after}."""
    buf = io.BytesIO()
    half_inch = 0.5 * inch
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=half_inch, rightMargin=half_inch,
        topMargin=half_inch, bottomMargin=half_inch,
    )
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle(
        'adv_title', parent=styles['Title'], fontSize=18, leading=22, alignment=1)
    sub_style = ParagraphStyle(
        'adv_sub', parent=styles['Normal'], fontSize=10, leading=14, alignment=1,
        textColor=colors.HexColor('#64748B'))

    story.append(Paragraph("Advance Salary Statement", title_style))
    meta = f"Officer: <b>{officer_name or '—'}</b>"
    if client_name:
        meta += f" &nbsp;·&nbsp; Client: <b>{client_name}</b>"
    story.append(Paragraph(meta, sub_style))
    story.append(Paragraph(
        f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}", sub_style))
    story.append(Spacer(1, 0.4 * cm))

    label_style = ParagraphStyle(
        'adv_cell', parent=styles['Normal'], fontSize=9, leading=12)
    money_style = ParagraphStyle(
        'adv_money', parent=styles['Normal'], fontSize=9, leading=12, alignment=2)
    head_style = ParagraphStyle(
        'adv_head', parent=styles['Normal'], fontSize=9, leading=12,
        textColor=colors.white)
    head_money = ParagraphStyle(
        'adv_head_money', parent=styles['Normal'], fontSize=9, leading=12,
        textColor=colors.white, alignment=2)

    table_data = [[
        Paragraph("<b>Date</b>", head_style),
        Paragraph("<b>Type</b>", head_style),
        Paragraph("<b>Note</b>", head_style),
        Paragraph("<b>Amount</b>", head_money),
        Paragraph("<b>Balance</b>", head_money),
    ]]
    for e in entries:
        etype = "Advance Taken" if e.get("type") == "advance" else "Repayment"
        amt = float(e.get("amount") or 0)
        amt_str = f"${amt:,.2f}" if e.get("type") == "advance" else f"-${amt:,.2f}"
        table_data.append([
            Paragraph(str(e.get("entry_date") or "—"), label_style),
            Paragraph(etype, label_style),
            Paragraph(str(e.get("note") or "—"), label_style),
            Paragraph(amt_str, money_style),
            Paragraph(f"${float(e.get('balance_after') or 0):,.2f}", money_style),
        ])
    if not entries:
        table_data.append([
            Paragraph("No advance transactions recorded.", label_style),
            Paragraph("", label_style), Paragraph("", label_style),
            Paragraph("", money_style), Paragraph("", money_style),
        ])

    tbl = Table(table_data, colWidths=[2.6 * cm, 3.2 * cm, 6.5 * cm, 3.0 * cm, 3.0 * cm])
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#E2E8F0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC')))
    tbl.setStyle(TableStyle(style))
    story.append(tbl)
    story.append(Spacer(1, 0.5 * cm))

    totals = [
        [Paragraph("Total Advanced", label_style),
         Paragraph(f"${float(total_advanced or 0):,.2f}", money_style)],
        [Paragraph("Total Repaid", label_style),
         Paragraph(f"-${float(total_repaid or 0):,.2f}", money_style)],
        [Paragraph("<b>Remaining Advance Balance</b>", label_style),
         Paragraph(f"<b>${float(remaining_balance or 0):,.2f}</b>", money_style)],
    ]
    totals_tbl = Table(totals, colWidths=[6.0 * cm, 4.0 * cm], hAlign='RIGHT')
    totals_tbl.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#CBD5E1')),
        ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#FBE4EA')),
        ('BOX', (0, 2), (-1, 2), 1.2, colors.HexColor('#0F172A')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(totals_tbl)

    doc.build(story)
    return buf.getvalue()



def build_csv(rows: list, columns: list) -> bytes:
    """columns: list of (header, key)"""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([h for h, _ in columns])
    for r in rows:
        w.writerow([r.get(k, "") if r.get(k) is not None else "" for _, k in columns])
    return buf.getvalue().encode("utf-8")


def build_xlsx(rows: list, columns: list, *, title: str = "Report",
               subtitle: str = "") -> bytes:
    """Excel workbook that mirrors the PDF report formatting: indigo header
    band, banded rows, red-bold post pin, currency-aware widths. Preserves
    the visual identity users see on-screen so exports feel like a designed
    report rather than raw CSV."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Report")[:31] or "Sheet1"

    indigo = "4F46E5"
    band_a = "FFFFFF"
    band_b = "F8FAFC"
    grid = "E2E8F0"
    text_dark = "0F172A"
    red = "DC2626"

    thin = Side(style="thin", color=grid)
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Title + subtitle rows on top
    ncols = max(len(columns), 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tcell = ws.cell(row=1, column=1, value=title)
    tcell.font = Font(name="Calibri", size=16, bold=True, color=text_dark)
    tcell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24

    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        scell = ws.cell(row=2, column=1, value=subtitle)
        scell.font = Font(name="Calibri", size=10, color="64748B")

    header_row_idx = 3 if subtitle else 2

    # Header
    for c, (label, _key) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row_idx, column=c, value=label)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=indigo)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = border

    ws.row_dimensions[header_row_idx].height = 22

    # Body
    money_keys = {
        "duty_rate", "billing_rate", "hourly_rate", "total", "cost_amount",
        "billing_amount", "margin"
    }

    pin_keys = {"post_pin", "post_pin_display"}

    numeric_keys = money_keys | {
        "duty_hours", "total_hours", "actual_hours",
        "completed_hours", "total_shifts", "completed",
        "absent", "late", "early_checkout", "cancelled",
        "confirmed", "attendance_pct", "coverage_pct"
    }

    for r, row in enumerate(rows, start=1):
        excel_row = header_row_idx + r
        band = band_a if r % 2 == 1 else band_b

        for c, (_label, key) in enumerate(columns, start=1):
            val = row.get(key)

            if val is None:
                val = ""

            cell = ws.cell(row=excel_row, column=c, value=val)
            cell.border = border
            cell.fill = PatternFill("solid", fgColor=band)

            if key in money_keys and isinstance(val, (int, float)):
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(
                    horizontal="right",
                    vertical="center"
                )

            elif key in numeric_keys and isinstance(val, (int, float)):
                cell.alignment = Alignment(
                    horizontal="right",
                    vertical="center"
                )

            else:
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=False
                )

            if key in pin_keys and val:
                cell.font = Font(
                    name="Calibri",
                    size=11,
                    bold=True,
                    color=red
                )
            else:
                cell.font = Font(
                    name="Calibri",
                    size=11,
                    color=text_dark
                )

    # Column widths — a bit larger for text-heavy columns
    width_by_key = {
        "date": 12,
        "shift_type": 12,
        "start_time": 10,
        "end_time": 10,
        "duty_hours": 10,
        "hourly_rate": 12,
        "total": 12,
        "duty_rate": 12,
        "billing_rate": 12,
        "cost_amount": 12,
        "billing_amount": 12,
        "margin": 12,
        "attendance_pct": 12,
        "coverage_pct": 12,
        "total_shifts": 10,
        "completed": 12,
        "absent": 10,
        "late": 10,
        "early_checkout": 12,
        "confirmed": 12,
        "cancelled": 12,
        "work_order_number": 14,
        "post_pin": 12,
        "post_pin_display": 18,
        "post_site_name": 26,
        "post_site_address": 26,
        "city": 14,
        "client_name": 22,
        "vendor_name": 22,
        "officer_name": 22,
        "confirmation_status": 16,
        "confirmation_method": 16,
        "shift_status": 14,
        "remarks": 32,
        "last_modified_by_name": 20,
        "last_modified_action": 18,
        "actual_check_in": 14,
        "actual_check_out": 14,
    }

    for c, (_, key) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(c)].width = width_by_key.get(key, 14)

    # Freeze the header
    ws.freeze_panes = ws.cell(
        row=header_row_idx + 1,
        column=1
    ).coordinate

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_pdf(title: str, subtitle: str, rows: list, columns: list) -> bytes:
    """Simple tabular PDF with wrapped and centered cell text."""
    buf = io.BytesIO()

    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm
    )

    styles = getSampleStyleSheet()
    story = []

    story.append(
        Paragraph(
            title,
            ParagraphStyle(
                't',
                parent=styles['Title'],
                fontSize=16
            )
        )
    )

    story.append(
        Paragraph(
            subtitle,
            ParagraphStyle(
                's',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.grey
            )
        )
    )

    story.append(Spacer(1, 0.4 * cm))

    header = [h for h, _ in columns]

    # Wrap body text inside cells
    cell_style = ParagraphStyle(
        'pdf_cell',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        alignment=1,
        textColor=colors.HexColor('#0F172A'),
    )

    header_style = ParagraphStyle(
        'pdf_header',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        alignment=1,
        textColor=colors.white,
    )

    body = []

    for r in rows:
        row_data = []

        for _, k in columns:
            value = r.get(k)

            if value is None:
                value = ""

            row_data.append(
                Paragraph(
                    str(value),
                    cell_style
                )
            )

        body.append(row_data)

    header_paragraphs = [
        Paragraph(str(h), header_style)
        for h in header
    ]

    tbl = Table(
        [header_paragraphs] + body,
        repeatRows=1
    )

    tbl.setStyle(
        TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
                colors.white,
                colors.HexColor('#F8FAFC')
            ]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ])
    )

    story.append(tbl)

    story.append(Spacer(1, 0.5 * cm))

    story.append(
        Paragraph(
            f"Generated {_dhaka_now_str()} (Asia/Dhaka)",
            ParagraphStyle(
                'f',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.grey
            )
        )
    )

    doc.build(story)
    return buf.getvalue()


# ---- Branded per-officer payslip PDF (matches user's mockup) ---------------

def _fmt_date(iso: str) -> str:
    """'2026-08-01' → 'Sat, 1 Aug' (short, weekday-first)."""
    try:
        d = datetime.strptime(iso, "%Y-%m-%d")
        return d.strftime("%a, %d %b")
    except Exception:
        return iso or ""


def _fmt_date_long(iso: str) -> str:
    try:
        d = datetime.strptime(iso, "%Y-%m-%d")

        # e.g. "31st July 2026"
        day = d.day
        suffix = (
            "th"
            if 4 <= day % 100 <= 20
            else {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
        )

        return f"{day}{suffix} {d.strftime('%B %Y')}"
    except Exception:
        return iso or ""


def _resolve_logo_bytes(logo_url: str) -> bytes | None:
    """Given a URL like '/api/files/officeflow/...' or a legacy absolute URL,
    resolve to the on-disk file inside STORAGE_ROOT so the PDF can embed
    the image."""
    if not logo_url:
        return None

    marker = "/api/files/"
    rel = None

    if marker in logo_url:
        rel = logo_url.split(marker, 1)[1]

    elif not logo_url.startswith(("http://", "https://")):
        rel = logo_url.lstrip("/")

    if not rel:
        return None

    root = os.environ.get(
        "STORAGE_ROOT",
        "/app/backend/uploads"
    )

    path = os.path.join(root, rel)

    try:
        with open(path, "rb") as fh:
            return fh.read()
    except Exception:
        return None


def build_officer_payslip_pdf(
    *,
    client_name: str | None,
    client_logo_url: str | None,
    officer_name: str,
    date_from: str,
    date_to: str,
    rows: list,
    show_financial: bool,
    extra_payment_rows: list = None,
    deduction_rows: list = None,
    advance_taken: float = 0.0,
    advance_repaid: float = 0.0,
    remaining_balance: float = 0.0,
    advance_entries: list = None,
    client_address: str | None = None,
    client_phone: str | None = None,
    client_email: str | None = None,
    client_website: str | None = None,
    officer_ssn: str | None = None,
) -> bytes:
    """Branded per-officer/per-client payslip in the layout the customer asked
    for (Arseas Security Service mockup).

    Columns: Date · Shift · Start · End · Duty Hours · Hourly Rate · Total ·
             Post Site · City · Post Site Pin · Remarks
    Footer: Total Duty Hours · Total Amount
    """

    buf = io.BytesIO()

    half_inch = 0.5 * inch

    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=half_inch,
        rightMargin=half_inch,
        topMargin=half_inch,
        bottomMargin=half_inch,
    )

    styles = getSampleStyleSheet()
    story = []

    # --- Header block: Logo + Client Title, both CENTERED ------------------

    logo_bytes = _resolve_logo_bytes(client_logo_url)
    center_cells = []

    if logo_bytes:
        try:
            from PIL import Image as PILImage

            with PILImage.open(io.BytesIO(logo_bytes)) as im:
                im.verify()

            center_cells.append(
                Image(
                    io.BytesIO(logo_bytes),
                    width=2.4 * cm,
                    height=2.4 * cm,
                    kind='proportional'
                )
            )
        except Exception:
            pass

    center_cells.append(
        Paragraph(
            f"<para align='center'><b>{client_name or '&nbsp;'}</b></para>",
            ParagraphStyle(
                'ct',
                parent=styles['Title'],
                fontSize=20,
                leading=24,
                alignment=1
            ),
        )
    )

    # Centered contact block: address / phone / email / website — each on its
    # own line, only rendered when the field is populated on the client doc.
    contact_style = ParagraphStyle(
        'contact', parent=styles['Normal'],
        fontSize=9, leading=12, alignment=1,
        textColor=colors.HexColor('#334155'),
    )
    if client_address:
        center_cells.append(
            Paragraph(f"<para align='center'>{client_address}</para>", contact_style)
        )
    contact_bits = []
    if client_phone:
        contact_bits.append(f"Tel: {client_phone}")
    if client_email:
        contact_bits.append(f"Email: {client_email}")
    if contact_bits:
        center_cells.append(
            Paragraph(
                f"<para align='center'>{' &nbsp;·&nbsp; '.join(contact_bits)}</para>",
                contact_style,
            )
        )
    if client_website:
        center_cells.append(
            Paragraph(
                f"<para align='center'>Website: {client_website}</para>",
                contact_style,
            )
        )

    center_stack = Table(
        [[c] for c in center_cells],
        colWidths=[doc.width]
    )

    center_stack.setStyle(
        TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
        ])
    )

    # Officer + Duty Periods block

    meta_style = ParagraphStyle(
        'meta',
        parent=styles['Normal'],
        fontSize=10,
        alignment=0
    )

    meta = [
        [
            Paragraph("<b>Security Officer's Name:</b>", meta_style),
            Paragraph(officer_name or "—", meta_style)
        ],
        [
            Paragraph("<b>Social Security Code:</b>", meta_style),
            Paragraph(officer_ssn or "—", meta_style)
        ],
        [
            Paragraph("<b>Duty Periods:</b>", meta_style),
            Paragraph(
                f"{_fmt_date_long(date_from)} to {_fmt_date_long(date_to)}",
                meta_style
            )
        ],
    ]

    meta_tbl = Table(
        meta,
        colWidths=[4.5 * cm, doc.width - 4.5 * cm],
        hAlign='LEFT'
    )

    meta_tbl.setStyle(
        TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ])
    )

    story.append(center_stack)
    story.append(Spacer(1, 0.3 * cm))
    story.append(meta_tbl)
    story.append(Spacer(1, 0.4 * cm))

    # --- Table --------------------------------------------------------------

    columns = [
        ("Date", "date"),
        ("Shift", "shift_type"),
        ("Start Time", "start_time"),
        ("End Time", "end_time"),
        ("Duty Hours", "duty_hours"),
        ("Hourly Rate", "hourly_rate"),
        ("Total", "total"),
        ("Post Site Name", "post_site_name"),
        ("City", "city"),
        ("Post Site Pin", "post_pin_display"),
        ("Remarks", "remarks"),
    ]

    if not show_financial:
        columns = [
            c for c in columns
            if c[1] not in ("hourly_rate", "total")
        ]

    header_labels = [h for h, _ in columns]

    body = []
    total_duty = 0.0
    total_amount = 0.0

    for r in rows:
        line = []

        for _, k in columns:
            v = r.get(k)

            if v is None:
                line.append("")

            elif k in ("hourly_rate", "total"):
                line.append(
                    f"${float(v):,.2f}" if v else ""
                )

            elif k == "duty_hours":
                line.append(
                    f"{float(v):g}"
                    if isinstance(v, (int, float)) and v
                    else str(v)
                )

            elif k == "date":
                line.append(_fmt_date(v))

            else:
                line.append(str(v))

        body.append(line)

        total_duty += float(r.get("duty_hours") or 0)
        total_amount += float(r.get("total") or 0)

    # Footer totals row

    footer = [""] * len(columns)

    def _colidx(key):
        for i, (_, k) in enumerate(columns):
            if k == key:
                return i
        return -1

    fi_duty = _colidx("duty_hours")
    fi_total = _colidx("total")

    if fi_duty >= 0:
        footer[fi_duty] = f"{total_duty:g}"

    if fi_total >= 0:
        footer[fi_total] = f"${total_amount:,.2f}"

    data = [header_labels] + body + [footer]

    weights_by_key = {
        "date": 1.4,
        "shift_type": 0.9,
        "start_time": 0.9,
        "end_time": 0.9,
        "duty_hours": 0.9,
        "hourly_rate": 1.0,
        "total": 1.0,
        "post_site_name": 2.6,
        "city": 1.1,
        "post_pin": 1.1,
        "post_pin_display": 1.4,
        "remarks": 2.2,
    }

    weights = [
        weights_by_key.get(k, 1.0)
        for _, k in columns
    ]

    total_w = sum(weights)

    col_widths = [
        doc.width * (w / total_w)
        for w in weights
    ]

    tbl = Table(
        data,
        colWidths=col_widths,
        repeatRows=1,
        hAlign='LEFT'
    )

    numeric_keys = {
        "duty_hours",
        "hourly_rate",
        "total"
    }

    numeric_col_idxs = [
        i
        for i, (_, k) in enumerate(columns)
        if k in numeric_keys
    ]

    left_keys = {
        "post_site_name",
        "remarks"
    }

    left_col_idxs = [
        i
        for i, (_, k) in enumerate(columns)
        if k in left_keys
    ]

    style = [
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F2C4CE')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),

        # Body defaults
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#94A3B8')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        # All cells centered
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),

        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),

        # Date column
        ('BACKGROUND', (0, 1), (0, -2), colors.HexColor('#FBE4EA')),

        # Footer totals row
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F8FAFC')),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.HexColor('#0F172A')),
    ]

    # Keep numeric columns centered as requested.
    for i in numeric_col_idxs:
        style.append(
            ('ALIGN', (i, 1), (i, -1), 'CENTER')
        )

    # Keep text-heavy columns centered as requested.
    for i in left_col_idxs:
        style.append(
            ('ALIGN', (i, 1), (i, -2), 'CENTER')
        )

    # Post Site Pin column — red bold text
    pin_idx = _colidx("post_pin_display")

    if pin_idx < 0:
        pin_idx = _colidx("post_pin")

    if pin_idx >= 0:
        style.append(
            (
                'TEXTCOLOR',
                (pin_idx, 1),
                (pin_idx, -2),
                colors.HexColor('#DC2626')
            )
        )

        style.append(
            (
                'FONTNAME',
                (pin_idx, 1),
                (pin_idx, -2),
                'Helvetica-Bold'
            )
        )

    tbl.setStyle(TableStyle(style))

    story.append(tbl)

    # --- Payslip summary: Gross / Extra / Deductions / Advance / Net --------

    if show_financial:
        gross = float(total_amount or 0)
        extra_rows = extra_payment_rows or []
        extra_total = sum(float(r.get("amount") or 0) for r in extra_rows)
        ded_rows = deduction_rows or []
        ded_total = sum(float(r.get("amount") or 0) for r in ded_rows)
        # Advances have zero effect on net pay.
        net_payment = gross + extra_total - ded_total

        money_style = ParagraphStyle(
            'payslip_money',
            parent=styles['Normal'],
            fontSize=10,
            leading=13,
            alignment=2,
        )

        label_style = ParagraphStyle(
            'payslip_money_label',
            parent=styles['Normal'],
            fontSize=9,
            leading=12,
            alignment=0,
        )

        final_label_style = ParagraphStyle(
            'payslip_final_label',
            parent=styles['Normal'],
            fontSize=11,
            leading=14,
            alignment=0,
        )

        final_money_style = ParagraphStyle(
            'payslip_final_money',
            parent=styles['Normal'],
            fontSize=14,
            leading=17,
            alignment=2,
        )

        summary_data = [
            [
                Paragraph("<b>Gross Pay</b>", label_style),
                Paragraph(f"<b>${gross:,.2f}</b>", money_style),
            ],
        ]

        if extra_rows:
            summary_data.append([
                Paragraph("<b>Extra Payments</b>", label_style),
                Paragraph("", money_style),
            ])
            for r in extra_rows:
                purpose = str(r.get("purpose") or "Extra Payment")
                date_v = str(r.get("date") or "")
                lbl = f"&nbsp;&nbsp;{purpose}"
                if date_v:
                    lbl += f" ({date_v})"
                summary_data.append([
                    Paragraph(lbl, label_style),
                    Paragraph(f"${float(r.get('amount') or 0):,.2f}", money_style),
                ])
        else:
            summary_data.append([
                Paragraph("Extra Payments", label_style),
                Paragraph("$0.00", money_style),
            ])

        # Manual deductions: itemized line items (styled like Extra Payments)
        # that reduce net pay. Fully independent from the advance ledger.
        if ded_rows:
            summary_data.append([
                Paragraph("<b>Deductions</b>", label_style),
                Paragraph("", money_style),
            ])
            for r in ded_rows:
                purpose = str(r.get("purpose") or "Deduction")
                date_v = str(r.get("date") or "")
                lbl = f"&nbsp;&nbsp;{purpose}"
                if date_v:
                    lbl += f" ({date_v})"
                summary_data.append([
                    Paragraph(lbl, label_style),
                    Paragraph(f"-${float(r.get('amount') or 0):,.2f}", money_style),
                ])
            summary_data.append([
                Paragraph("Deductions Total", label_style),
                Paragraph(f"-${ded_total:,.2f}", money_style),
            ])

        net_row_idx = len(summary_data)
        summary_data.append([
            Paragraph("<b>NET PAYMENT</b>", final_label_style),
            Paragraph(f"<b>${net_payment:,.2f}</b>", final_money_style),
        ])
        bal_row_idx = len(summary_data)
        summary_data.append([
            Paragraph("Remaining Advance Balance", label_style),
            Paragraph(f"${float(remaining_balance or 0):,.2f}", money_style),
        ])

        summary_tbl = Table(
            summary_data,
            colWidths=[6.5 * cm, 4.0 * cm],
            hAlign='RIGHT',
        )

        summary_tbl.setStyle(
            TableStyle([
                ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#CBD5E1')),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F8FAFC')),
                ('BACKGROUND', (0, net_row_idx), (-1, net_row_idx), colors.HexColor('#FBE4EA')),
                ('BOX', (0, net_row_idx), (-1, net_row_idx), 1.2, colors.HexColor('#0F172A')),
                ('BACKGROUND', (0, bal_row_idx), (-1, bal_row_idx), colors.HexColor('#F8FAFC')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ])
        )

        story.append(Spacer(1, 0.5 * cm))
        story.append(summary_tbl)

    story.append(Spacer(1, 0.4 * cm))

    doc.build(story)

    return buf.getvalue()


# ---- Client → Vendor Invoice PDF (matches Arseas mockup) --------------------

def build_invoice_pdf(
    *,
    client: dict,
    vendor: dict,
    invoice_number: str,
    invoice_date: str,
    billing_period_from: str,
    billing_period_to: str,
    lines: list,
    total_hours: float,
    total_amount: float,
    amount_in_words: str = "",
    accent_color: str = None,
) -> bytes:
    """Portrait A4 invoice — black header/footer bars with an accent line
    (defaults to gold, overridable via ``accent_color`` for per-client brand),
    faint centered watermark logo, invoice metadata block and centered
    wrapped table cells.
    """

    black = colors.HexColor('#000000')

    # Fall back to gold when the client has no brand color configured.
    def _safe_hex(val, fallback):
        try:
            return colors.HexColor(val) if val else fallback
        except Exception:
            return fallback

    default_gold = colors.HexColor('#D4A017')
    golden = _safe_hex(accent_color, default_gold)
    dark = colors.HexColor('#0F172A')

    # ---- Layout constants -----------------------------------------------

    PAGE_MARGIN = 0.4 * inch
    TOP_MARGIN = 0.2 * inch
    HEADER_TOP_GAP = 30
    HEADER_BAR_H = 40
    LOGO_OVERFLOW = 27
    HEADER_GOLD_H = 3
    HEADER_GOLD_FROM_BOTTOM = 1
    HEADER_GOLD_W_RATIO = 0.60
    FOOTER_BLACK_H = 24
    FOOTER_GOLD_H = 3
    FOOTER_CONTENT_GAP = 20

    buf = io.BytesIO()

    top_margin = (
        TOP_MARGIN
        + HEADER_TOP_GAP
        + HEADER_BAR_H
        + LOGO_OVERFLOW
    )

    bottom_margin = (
        FOOTER_BLACK_H
        + FOOTER_GOLD_H
        + FOOTER_CONTENT_GAP
    )

    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=PAGE_MARGIN,
        rightMargin=PAGE_MARGIN,
        topMargin=top_margin,
        bottomMargin=bottom_margin,
    )

    styles = getSampleStyleSheet()
    story = []

    logo_bytes = _resolve_logo_bytes(
        client.get('logo_url') or client.get('logo_path')
    )

    if logo_bytes:
        try:
            from PIL import Image as PILImage

            with PILImage.open(io.BytesIO(logo_bytes)) as im:
                im.verify()

        except Exception:
            logo_bytes = None

    def _draw_chrome(canvas, doc_):
        page_w, page_h = A4

        canvas.saveState()

        # ---- Watermark ---------------------------------------------------

        if logo_bytes:
            try:
                from reportlab.lib.utils import ImageReader

                img = ImageReader(io.BytesIO(logo_bytes))
                iw, ih = img.getSize()

                target_w = 4.5 * inch
                scale = target_w / float(iw)
                target_h = ih * scale

                cx = (page_w - target_w) / 2

                cy = (
                    page_h * 0.35
                    - target_h / 2
                )

                canvas.saveState()

                canvas.setFillAlpha(0.10)

                canvas.drawImage(
                    img,
                    cx,
                    cy,
                    width=target_w,
                    height=target_h,
                    mask='auto',
                    preserveAspectRatio=True
                )

                canvas.restoreState()

            except Exception:
                pass

        # ---- Header black bar -------------------------------------------

        bar_top = (
            page_h
            - TOP_MARGIN
            - HEADER_TOP_GAP
        )

        bar_bottom = (
            bar_top
            - HEADER_BAR_H
        )

        canvas.setFillColor(black)

        canvas.rect(
            PAGE_MARGIN,
            bar_bottom,
            page_w - 2 * PAGE_MARGIN,
            HEADER_BAR_H,
            fill=1,
            stroke=0
        )

        # ---- Golden accent line -----------------------------------------

        gold_w = page_w * HEADER_GOLD_W_RATIO
        gold_x = PAGE_MARGIN
        gold_y = (
            bar_bottom
            + HEADER_GOLD_FROM_BOTTOM
        )

        canvas.setFillColor(golden)

        canvas.rect(
            gold_x,
            gold_y,
            gold_w,
            HEADER_GOLD_H,
            fill=1,
            stroke=0
        )

        # ---- Header logo ------------------------------------------------

        content_w = page_w - 2 * PAGE_MARGIN

        logo_area_w = content_w * 0.60
        name_area_w = content_w * 0.40

        if logo_bytes:
            try:
                from reportlab.lib.utils import ImageReader

                img = ImageReader(
                    io.BytesIO(logo_bytes)
                )

                iw, ih = img.getSize()

                logo_h_target = (
                    HEADER_BAR_H
                    + 2 * LOGO_OVERFLOW
                )

                scale_h = (
                    logo_h_target
                    / float(ih)
                )

                max_lw = logo_area_w - 16

                scale_w = (
                    max_lw
                    / float(iw)
                )

                scale = min(
                    scale_h,
                    scale_w
                )

                lw = iw * scale
                logo_h = ih * scale

                lx = PAGE_MARGIN + 8

                ly = (
                    bar_bottom
                    + (HEADER_BAR_H - logo_h) / 2
                )

                canvas.drawImage(
                    img,
                    lx,
                    ly,
                    width=lw,
                    height=logo_h,
                    mask='auto',
                    preserveAspectRatio=True
                )

            except Exception:
                pass

        # ---- Client name inside header ---------------------------------

        cname = (
            client.get('name') or ''
        ).upper()

        if cname:
            from reportlab.platypus import Paragraph as _P

            for size in (14, 12, 10, 9):

                cname_style = ParagraphStyle(
                    'cname',
                    fontName='Helvetica-Bold',
                    fontSize=size,
                    leading=size + 3,
                    textColor=colors.white,
                    alignment=2,
                )

                p = _P(
                    cname,
                    cname_style
                )

                max_w = name_area_w - 8

                _w, _h = p.wrap(
                    max_w,
                    HEADER_BAR_H
                )

                if _h <= HEADER_BAR_H - 4:
                    break

            p.drawOn(
                canvas,
                page_w - PAGE_MARGIN - max_w - 4,
                bar_bottom
                + (HEADER_BAR_H - _h) / 2
            )

        # ---- Footer -----------------------------------------------------

        bar_w = page_w * 0.95

        bar_x = page_w - bar_w

        canvas.setFillColor(black)

        canvas.rect(
            bar_x,
            0,
            bar_w,
            FOOTER_BLACK_H,
            fill=1,
            stroke=0
        )

        canvas.setFillColor(golden)

        canvas.rect(
            bar_x,
            FOOTER_BLACK_H,
            bar_w,
            FOOTER_GOLD_H,
            fill=1,
            stroke=0
        )

        canvas.setFillColor(colors.white)

        canvas.setFont(
            'Helvetica',
            13
        )

        text_y = (
            (FOOTER_BLACK_H - 13) / 2
            + 3
        )

        canvas.drawCentredString(
            page_w / 2,
            text_y,
            'Thanks for business with us!'
        )

        canvas.restoreState()

    # ---- INVOICE title ---------------------------------------------------

    story.append(
        Paragraph(
            "<para align='center'><b>INVOICE</b></para>",
            ParagraphStyle(
                'title',
                parent=styles['Title'],
                fontSize=22,
                leading=26,
                textColor=dark
            ),
        )
    )

    story.append(
        Spacer(1, 0.35 * cm)
    )

    # ---- BILLING TO / BILLING FROM --------------------------------------

    label_style_left = ParagraphStyle(
        'lbl_l',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.white,
        spaceAfter=0,
        leading=12,
        alignment=0,
    )

    label_style_right = ParagraphStyle(
        'lbl_r',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.white,
        spaceAfter=0,
        leading=12,
        alignment=2,
    )

    name_style_l = ParagraphStyle(
        'nm_l',
        parent=styles['Normal'],
        fontSize=10,
        textColor=dark,
        spaceAfter=1,
        alignment=0
    )

    name_style_r = ParagraphStyle(
        'nm_r',
        parent=styles['Normal'],
        fontSize=10,
        textColor=dark,
        spaceAfter=1,
        alignment=2
    )

    body_style_l = ParagraphStyle(
        'b_l',
        parent=styles['Normal'],
        fontSize=9,
        textColor=dark,
        leading=12,
        alignment=0
    )

    body_style_r = ParagraphStyle(
        'b_r',
        parent=styles['Normal'],
        fontSize=9,
        textColor=dark,
        leading=12,
        alignment=2
    )

    def label_tag(
        text: str,
        right_aligned: bool
    ) -> Table:

        p = Paragraph(
            f"<b>{text}</b>",
            label_style_right
            if right_aligned
            else label_style_left
        )

        pill = Table(
            [[p]],
            colWidths=[4.6 * cm]
        )

        pill.setStyle(
            TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), black),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                (
                    'ALIGN',
                    (0, 0),
                    (-1, -1),
                    'RIGHT'
                    if right_aligned
                    else 'LEFT'
                ),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ])
        )

        col_w = doc.width / 2 - 2
        pad_w = col_w - 4.6 * cm

        if right_aligned:
            wrapper = Table(
                [["", pill]],
                colWidths=[
                    pad_w,
                    4.6 * cm
                ]
            )
        else:
            wrapper = Table(
                [[pill, ""]],
                colWidths=[
                    4.6 * cm,
                    pad_w
                ]
            )

        wrapper.setStyle(
            TableStyle([
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ])
        )

        return wrapper

    def contact_block(
        label: str,
        party: dict,
        right_aligned: bool
    ) -> Table:

        nstyle = (
            name_style_r
            if right_aligned
            else name_style_l
        )

        bstyle = (
            body_style_r
            if right_aligned
            else body_style_l
        )

        details = []

        phone = (
            party.get('phone')
            or party.get('contact_number')
            or party.get('contact_phone')
        )

        if party.get('address'):
            details.append(
                party['address']
            )

        if party.get('city'):
            details.append(
                party['city']
            )

        if phone:
            details.append(
                f"Phone: {phone}"
            )

        if party.get('email'):
            details.append(
                f"Email: {party['email']}"
            )

        if party.get('website'):
            details.append(
                f"Web: {party['website']}"
            )

        cells = [
            [
                label_tag(
                    label,
                    right_aligned
                )
            ],
            [
                Spacer(
                    1,
                    0.15 * cm
                )
            ],
            [
                Paragraph(
                    f"<b>{party.get('name', '')}</b>",
                    nstyle
                )
            ],
        ]

        for d in details:
            cells.append(
                [
                    Paragraph(
                        d,
                        bstyle
                    )
                ]
            )

        t = Table(
            cells,
            colWidths=[
                doc.width / 2 - 2
            ],
            hAlign=(
                'RIGHT'
                if right_aligned
                else 'LEFT'
            )
        )

        t.setStyle(
            TableStyle([
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 1),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
            ])
        )

        return t

    def meta_block() -> Table:

        rows = [
            [
                Paragraph(
                    "<b>Invoice No:</b>",
                    body_style_r
                ),
                Paragraph(
                    invoice_number or "—",
                    body_style_r
                )
            ],
            [
                Paragraph(
                    "<b>Invoice Date:</b>",
                    body_style_r
                ),
                Paragraph(
                    invoice_date or "—",
                    body_style_r
                )
            ],
            [
                Paragraph(
                    "<b>Billing Period:</b>",
                    body_style_r
                ),
                Paragraph(
                    f"{billing_period_from} to {billing_period_to}",
                    body_style_r
                )
            ],
        ]

        t = Table(
            rows,
            colWidths=[
                3.4 * cm,
                5.1 * cm
            ],
            hAlign='RIGHT'
        )

        t.setStyle(
            TableStyle([
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 1),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ])
        )

        return t

    parties = Table(
        [
            [
                contact_block(
                    "BILLING FROM",
                    client,
                    right_aligned=False
                ),
                contact_block(
                    "BILLING TO",
                    vendor,
                    right_aligned=True
                )
            ],
            [
                Spacer(1, 0.15 * cm),
                Spacer(1, 0.15 * cm)
            ],
            [
                "",
                meta_block()
            ],
        ],
        colWidths=[
            doc.width / 2,
            doc.width / 2
        ]
    )

    parties.setStyle(
        TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('ALIGN', (1, 2), (1, 2), 'RIGHT'),
        ])
    )

    story.append(parties)

    story.append(
        Spacer(1, 0.4 * cm)
    )

    # ---- Line-item table --------------------------------------------------
    #
    # IMPORTANT:
    # Every table cell is now a Paragraph.
    # This allows ReportLab to wrap long text automatically inside the
    # existing cell width instead of allowing it to overflow outside.
    #
    # All paragraph alignments are CENTER, so Work Order, Location,
    # Shift Date, Actual Hour, Rate and Total Amount are all centered.

    header_row = [
        "Shift Date",
        "Location",
        "Work Order",
        "Actual Hour",
        "Rate",
        "Total Amount"
    ]

    # Centered wrapped text styles
    invoice_cell_style = ParagraphStyle(
        'invoice_cell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=11,
        textColor=dark,
        alignment=1,          # CENTER
        wordWrap='CJK',
        splitLongWords=True,
    )

    invoice_header_style = ParagraphStyle(
        'invoice_header',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=golden,
        alignment=1,          # CENTER
        wordWrap='CJK',
        splitLongWords=True,
    )

    invoice_footer_style = ParagraphStyle(
        'invoice_footer',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=golden,
        alignment=1,          # CENTER
        wordWrap='CJK',
        splitLongWords=True,
    )

    # Header itself is also Paragraph-based, so long header labels can wrap.
    header_cells = [
        Paragraph(
            str(value),
            invoice_header_style
        )
        for value in header_row
    ]

    body_rows = []

    for ln in lines:

        rate = ln.get("rate")
        total = ln.get("total_amount")

        shift_date = (
            ln.get("shift_date")
            or "—"
        )

        location = (
            ln.get("location")
            or "—"
        )

        work_order = (
            ln.get("work_order")
            or "—"
        )

        actual_hour = (
            f"{float(ln.get('actual_hours') or 0):.2f}"
        )

        rate_text = (
            f"$ {float(rate):,.2f}"
            if rate is not None
            else "—"
        )

        total_text = (
            f"$ {float(total):,.2f}"
            if total is not None
            else "—"
        )

        # Every field is a Paragraph.
        # Therefore long Location / Work Order text wraps inside its cell.
        body_rows.append([
            Paragraph(
                str(shift_date),
                invoice_cell_style
            ),
            Paragraph(
                str(location),
                invoice_cell_style
            ),
            Paragraph(
                str(work_order),
                invoice_cell_style
            ),
            Paragraph(
                actual_hour,
                invoice_cell_style
            ),
            Paragraph(
                rate_text,
                invoice_cell_style
            ),
            Paragraph(
                total_text,
                invoice_cell_style
            ),
        ])

    if not body_rows:
        body_rows.append([
            Paragraph("—", invoice_cell_style),
            Paragraph(
                "No completed shifts in this period",
                invoice_cell_style
            ),
            Paragraph("—", invoice_cell_style),
            Paragraph("0.00", invoice_cell_style),
            Paragraph("—", invoice_cell_style),
            Paragraph("—", invoice_cell_style),
        ])

    # Footer totals row.
    footer_values = [
        "",
        "",
        "Total",
        f"{float(total_hours):.2f}",
        "",
        f"$ {float(total_amount):,.2f}"
    ]

    footer_row = [
        Paragraph(
            str(value),
            invoice_footer_style
        )
        for value in footer_values
    ]

    data = [
        header_cells
    ] + body_rows + [
        footer_row
    ]

    # Existing column widths remain unchanged.
    col_widths = [
        doc.width * 0.14,
        doc.width * 0.22,
        doc.width * 0.24,
        doc.width * 0.12,
        doc.width * 0.13,
        doc.width * 0.15,
    ]

    tbl = Table(
        data,
        colWidths=col_widths,
        repeatRows=1,
        hAlign='LEFT'
    )

    style = [
        # Header row — black bg, bold golden text
        ('BACKGROUND', (0, 0), (-1, 0), black),
        ('TEXTCOLOR', (0, 0), (-1, 0), golden),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),

        # Body defaults
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 1), (-1, -2), dark),

        # Golden borders on all cells
        ('GRID', (0, 0), (-1, -1), 0.75, golden),

        # Vertically centered
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        # ALL CONTENT CENTERED
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),

        # Cell padding
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),

        # Totals row — same as header row
        ('BACKGROUND', (0, -1), (-1, -1), black),
        ('TEXTCOLOR', (0, -1), (-1, -1), golden),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),

        # Explicitly center totals row
        ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
    ]

    tbl.setStyle(
        TableStyle(style)
    )

    story.append(tbl)

    story.append(
        Spacer(1, 0.35 * cm)
    )

    if amount_in_words:
        story.append(
            Paragraph(
                f"<b>In-Words:</b> {amount_in_words}",
                ParagraphStyle(
                    'iw',
                    parent=styles['Normal'],
                    fontSize=9,
                    textColor=dark
                ),
            )
        )

    doc.build(
        story,
        onFirstPage=_draw_chrome,
        onLaterPages=_draw_chrome
    )

    return buf.getvalue()


# ---- Security Officer Payment (SO) report — matches payslip style ----------

def build_so_payment_report_pdf(*, ctx: dict, methods: list) -> bytes:
    """Per-officer payment transaction report.

    Table columns: Date · Payment Method · Transaction ID · Amount, with a
    Total footer row. Followed by a summary box (Client, Officer, Code,
    Statement Period, Total Balance and per-method balances). Visual style
    mirrors the Security Officer payslip (centered title, pink header band,
    pink date column, bold total footer, boxed summary).
    """
    buf = io.BytesIO()
    half_inch = 0.5 * inch
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=half_inch, rightMargin=half_inch,
        topMargin=half_inch, bottomMargin=half_inch,
    )
    styles = getSampleStyleSheet()
    story = []

    client_name = ctx.get("client_name")
    officer_name = ctx.get("officer_name")
    officer_code = ctx.get("officer_code")
    payments = ctx.get("payments") or []
    total = float(ctx.get("total") or 0)
    by_method = ctx.get("by_method") or {}
    period = f"{ctx.get('statement_from') or '—'} to {ctx.get('statement_to') or '—'}"

    # Centered client title
    story.append(Paragraph(
        f"<para align='center'><b>{client_name or '&nbsp;'}</b></para>",
        ParagraphStyle('ct', parent=styles['Title'], fontSize=20, leading=24, alignment=1)))
    story.append(Paragraph(
        "<para align='center'>Security Officer Payment Report</para>",
        ParagraphStyle('cst', parent=styles['Normal'], fontSize=11, leading=15,
                       alignment=1, textColor=colors.HexColor('#64748B'))))
    story.append(Spacer(1, 0.3 * cm))

    meta_style = ParagraphStyle('meta', parent=styles['Normal'], fontSize=10, alignment=0)
    meta = [
        [Paragraph("<b>Security Officer's Name:</b>", meta_style),
         Paragraph(officer_name or "—", meta_style)],
        [Paragraph("<b>Security Officer Code:</b>", meta_style),
         Paragraph(officer_code or "—", meta_style)],
        [Paragraph("<b>Statement Period:</b>", meta_style),
         Paragraph(period, meta_style)],
    ]
    meta_tbl = Table(meta, colWidths=[5.0 * cm, doc.width - 5.0 * cm], hAlign='LEFT')
    meta_tbl.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(meta_tbl)
    story.append(Spacer(1, 0.4 * cm))

    # --- Transactions table ---
    header = ["Date", "Payment Method", "Transaction ID", "Amount"]
    body = []
    for p in payments:
        body.append([
            _fmt_date(p.get("date")),
            str(p.get("payment_method") or "—"),
            str(p.get("transaction_id") or "—"),
            f"${float(p.get('amount') or 0):,.2f}",
        ])
    if not body:
        body.append(["—", "No payments recorded", "—", "$0.00"])
    footer = ["", "", "Total", f"${total:,.2f}"]
    data = [header] + body + [footer]

    col_widths = [doc.width * 0.20, doc.width * 0.28, doc.width * 0.32, doc.width * 0.20]
    tbl = Table(data, colWidths=col_widths, repeatRows=1, hAlign='LEFT')
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F2C4CE')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#94A3B8')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('BACKGROUND', (0, 1), (0, -2), colors.HexColor('#FBE4EA')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F8FAFC')),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.HexColor('#0F172A')),
        ('ALIGN', (2, -1), (2, -1), 'RIGHT'),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 0.5 * cm))

    # --- Summary box ---
    lbl = ParagraphStyle('sum_lbl', parent=styles['Normal'], fontSize=9, leading=12, alignment=0)
    val = ParagraphStyle('sum_val', parent=styles['Normal'], fontSize=9, leading=12, alignment=2)
    final_lbl = ParagraphStyle('sum_flbl', parent=styles['Normal'], fontSize=11, leading=14, alignment=0)
    final_val = ParagraphStyle('sum_fval', parent=styles['Normal'], fontSize=13, leading=16, alignment=2)

    summary_data = [
        [Paragraph("<b>Client Name</b>", lbl), Paragraph(client_name or "—", val)],
        [Paragraph("<b>Officer Name</b>", lbl), Paragraph(officer_name or "—", val)],
        [Paragraph("<b>Officer Code</b>", lbl), Paragraph(officer_code or "—", val)],
        [Paragraph("<b>Statement Period</b>", lbl), Paragraph(period, val)],
    ]
    total_row_idx = len(summary_data)
    summary_data.append([
        Paragraph("<b>TOTAL BALANCE</b>", final_lbl),
        Paragraph(f"<b>${total:,.2f}</b>", final_val),
    ])
    # Per-method balances (only methods with a non-zero total)
    method_rows_idx = []
    for m in methods:
        amt = float(by_method.get(m, 0) or 0)
        if amt:
            method_rows_idx.append(len(summary_data))
            summary_data.append([
                Paragraph(f"&nbsp;&nbsp;Paid via {m}", lbl),
                Paragraph(f"${amt:,.2f}", val),
            ])

    summary_tbl = Table(summary_data, colWidths=[7.5 * cm, 4.0 * cm], hAlign='RIGHT')
    sstyle = [
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#CBD5E1')),
        ('BACKGROUND', (0, total_row_idx), (-1, total_row_idx), colors.HexColor('#FBE4EA')),
        ('BOX', (0, total_row_idx), (-1, total_row_idx), 1.2, colors.HexColor('#0F172A')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]
    for i in method_rows_idx:
        sstyle.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC')))
    summary_tbl.setStyle(TableStyle(sstyle))
    story.append(summary_tbl)

    story.append(Spacer(1, 0.4 * cm))
    story.append(Paragraph(
        f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        ParagraphStyle('f', parent=styles['Normal'], fontSize=8, textColor=colors.grey)))

    doc.build(story)
    return buf.getvalue()


def build_so_payment_report_xlsx(*, ctx: dict, methods: list) -> bytes:
    """Excel version of the SO payment report: transactions table + total row,
    then a summary block with per-method balances."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "SO Payments"

    pink = "F2C4CE"
    pink_soft = "FBE4EA"
    grid = "94A3B8"
    text_dark = "0F172A"
    thin = Side(style="thin", color="E2E8F0")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    client_name = ctx.get("client_name") or "—"
    officer_name = ctx.get("officer_name") or "—"
    officer_code = ctx.get("officer_code") or "—"
    payments = ctx.get("payments") or []
    total = float(ctx.get("total") or 0)
    by_method = ctx.get("by_method") or {}
    period = f"{ctx.get('statement_from') or '—'} to {ctx.get('statement_to') or '—'}"

    # Title
    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1, value=f"{client_name} — Security Officer Payment Report")
    c.font = Font(size=14, bold=True, color=text_dark)
    ws.row_dimensions[1].height = 22

    ws.cell(row=2, column=1, value="Officer Name").font = Font(bold=True)
    ws.cell(row=2, column=2, value=officer_name)
    ws.cell(row=3, column=1, value="Officer Code").font = Font(bold=True)
    ws.cell(row=3, column=2, value=officer_code)
    ws.cell(row=4, column=1, value="Statement Period").font = Font(bold=True)
    ws.cell(row=4, column=2, value=period)

    header_row = 6
    headers = ["Date", "Payment Method", "Transaction ID", "Amount"]
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = Font(bold=True, color=text_dark)
        cell.fill = PatternFill("solid", fgColor=pink)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    r = header_row
    for p in payments:
        r += 1
        vals = [p.get("date"), p.get("payment_method"),
                p.get("transaction_id") or "—", float(p.get("amount") or 0)]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.border = border
            if col == 4:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="center")
    if not payments:
        r += 1
        ws.cell(row=r, column=1, value="—").border = border
        ws.cell(row=r, column=2, value="No payments recorded").border = border
        ws.cell(row=r, column=3, value="—").border = border
        tc = ws.cell(row=r, column=4, value=0)
        tc.number_format = '"$"#,##0.00'
        tc.border = border

    # Total row
    r += 1
    tlabel = ws.cell(row=r, column=3, value="Total")
    tlabel.font = Font(bold=True)
    tlabel.alignment = Alignment(horizontal="right")
    tlabel.fill = PatternFill("solid", fgColor=pink_soft)
    tval = ws.cell(row=r, column=4, value=total)
    tval.font = Font(bold=True)
    tval.number_format = '"$"#,##0.00'
    tval.alignment = Alignment(horizontal="right")
    tval.fill = PatternFill("solid", fgColor=pink_soft)

    # Summary block
    r += 2
    ws.cell(row=r, column=1, value="Summary").font = Font(bold=True, size=12)
    summary = [("Client Name", client_name), ("Officer Name", officer_name),
               ("Officer Code", officer_code), ("Statement Period", period),
               ("Total Balance", total)]
    for label, value in summary:
        r += 1
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        vc = ws.cell(row=r, column=2, value=value)
        if label == "Total Balance":
            vc.number_format = '"$"#,##0.00'
            vc.font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value="Balance by Payment Method").font = Font(bold=True)
    for m in methods:
        amt = float(by_method.get(m, 0) or 0)
        if amt:
            r += 1
            ws.cell(row=r, column=1, value=f"Paid via {m}")
            mc = ws.cell(row=r, column=2, value=amt)
            mc.number_format = '"$"#,##0.00'

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 16

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---- Client Payment Records report (W2 / W9) -------------------------------

def _pay_cell(comp: dict) -> str:
    """Compact multi-line cell for a payment component (amount + date/txn)."""
    comp = comp or {}
    amt = float(comp.get("amount") or 0)
    parts = [f"${amt:,.2f}"]
    if comp.get("date"):
        parts.append(str(comp.get("date")))
    if comp.get("transaction_id"):
        parts.append(f"#{comp.get('transaction_id')}")
    return "<br/>".join(parts)


def _client_center_header(client: dict, doc, styles) -> list:
    """Centered statement header: logo, client name, then contact lines
    (Address · Email · Phone · Website)."""
    flow = []
    logo_bytes = _resolve_logo_bytes(client.get("logo_url") or client.get("logo_path"))
    if logo_bytes:
        try:
            from PIL import Image as PILImage
            with PILImage.open(io.BytesIO(logo_bytes)) as im:
                im.verify()
            img = Image(io.BytesIO(logo_bytes), width=2.4 * cm, height=2.4 * cm, kind='proportional')
            img.hAlign = 'CENTER'
            flow.append(img)
            flow.append(Spacer(1, 0.15 * cm))
        except Exception:
            pass
    flow.append(Paragraph(
        f"<para align='center'><b>{client.get('name') or ''}</b></para>",
        ParagraphStyle('cn', parent=styles['Title'], fontSize=18, leading=22, alignment=1)))
    contact_style = ParagraphStyle('cc', parent=styles['Normal'], fontSize=9, leading=13,
                                   alignment=1, textColor=colors.HexColor('#475569'))
    lines = []
    if client.get("address"):
        lines.append(str(client["address"]))
    row2 = " · ".join([x for x in [client.get("email"), client.get("contact_number")] if x])
    if row2:
        lines.append(row2)
    if client.get("website"):
        lines.append(str(client["website"]))
    for ln in lines:
        flow.append(Paragraph(f"<para align='center'>{ln}</para>", contact_style))
    return flow


def build_client_payment_records_pdf(*, ctx: dict) -> bytes:
    """Client statement: centered client header, a title/period row, then the
    payment table (SL, Officer Name, Address, Social Security, W2, grouped W9
    [Direct Deposit / Zelle Transfer / W9 Total], Total) with grand totals."""
    buf = io.BytesIO()
    half_inch = 0.5 * inch
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=half_inch, rightMargin=half_inch,
        topMargin=half_inch, bottomMargin=half_inch,
    )
    styles = getSampleStyleSheet()
    story = []

    client = ctx.get("client") or {}
    rows = ctx.get("rows") or []
    totals = ctx.get("totals") or {}
    period = ctx.get("period") or {}

    # Centered header
    story.extend(_client_center_header(client, doc, styles))
    story.append(Spacer(1, 0.4 * cm))

    # Title (left) + Statement Period (right)
    period_str = f"{period.get('from') or '—'} to {period.get('to') or '—'}"
    title_l = Paragraph("<b>Security Officer Payment Records</b>",
                        ParagraphStyle('tl', parent=styles['Normal'], fontSize=12, alignment=0))
    period_r = Paragraph(f"<para align='right'><b>Statement Period:</b> {period_str}</para>",
                         ParagraphStyle('pr', parent=styles['Normal'], fontSize=10, alignment=2,
                                        textColor=colors.HexColor('#475569')))
    title_row = Table([[title_l, period_r]], colWidths=[doc.width * 0.5, doc.width * 0.5])
    title_row.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0), ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(title_row)

    cell = ParagraphStyle('c', parent=styles['Normal'], fontSize=8, leading=10, alignment=1)
    cell_l = ParagraphStyle('cl', parent=styles['Normal'], fontSize=8, leading=10, alignment=0)
    head = ParagraphStyle('h', parent=styles['Normal'], fontSize=8, leading=10,
                          alignment=1, textColor=colors.HexColor('#0F172A'))

    def H(t):
        return Paragraph(f"<b>{t}</b>", head)

    header_row1 = [H("SL"), H("Security Officer Name"), H("Address"), H("Social Security"),
                   H("W2"), H("W9"), "", "", H("Total (W2+W9)")]
    header_row2 = ["", "", "", "", "", H("Direct Deposit"), H("Zelle Transfer"), H("W9 Total"), ""]
    data = [header_row1, header_row2]
    for r in rows:
        data.append([
            Paragraph(str(r.get("sl")), cell),
            Paragraph(str(r.get("officer_name") or "—"), cell_l),
            Paragraph(str(r.get("officer_address") or "—"), cell_l),
            Paragraph(str(r.get("social_security_code") or "—"), cell),
            Paragraph(f"${r.get('w2_amount', 0):,.2f}", cell),
            Paragraph(f"${r.get('w9_direct_deposit_amount', 0):,.2f}", cell),
            Paragraph(f"${r.get('w9_zelle_amount', 0):,.2f}", cell),
            Paragraph(f"${r.get('w9_total', 0):,.2f}", cell),
            Paragraph(f"<b>${r.get('total', 0):,.2f}</b>", cell),
        ])
    if not rows:
        data.append([Paragraph("—", cell), Paragraph("No payment records", cell_l)]
                    + [Paragraph("—", cell) for _ in range(7)])

    data.append([
        Paragraph("", cell), Paragraph("<b>Grand Total</b>", cell_l),
        Paragraph("", cell), Paragraph("", cell),
        Paragraph(f"<b>${totals.get('w2', 0):,.2f}</b>", cell),
        Paragraph(f"<b>${totals.get('w9_direct_deposit', 0):,.2f}</b>", cell),
        Paragraph(f"<b>${totals.get('w9_zelle', 0):,.2f}</b>", cell),
        Paragraph(f"<b>${totals.get('w9_total', 0):,.2f}</b>", cell),
        Paragraph(f"<b>${totals.get('grand_total', 0):,.2f}</b>", cell),
    ])

    weights = [0.5, 2.2, 2.6, 1.6, 1.6, 1.8, 1.8, 1.3, 1.5]
    total_w = sum(weights)
    col_widths = [doc.width * (w / total_w) for w in weights]

    tbl = Table(data, colWidths=col_widths, repeatRows=2, hAlign='LEFT')
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 1), colors.HexColor('#F2C4CE')),
        ('SPAN', (0, 0), (0, 1)),   # SL
        ('SPAN', (1, 0), (1, 1)),   # Name
        ('SPAN', (2, 0), (2, 1)),   # Address
        ('SPAN', (3, 0), (3, 1)),   # Social Security
        ('SPAN', (4, 0), (4, 1)),   # W2
        ('SPAN', (5, 0), (7, 0)),   # W9 group
        ('SPAN', (8, 0), (8, 1)),   # Total
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#94A3B8')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F8FAFC')),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.HexColor('#0F172A')),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 0.4 * cm))
    story.append(Paragraph(
        f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        ParagraphStyle('f', parent=styles['Normal'], fontSize=8, textColor=colors.grey)))

    doc.build(story)
    return buf.getvalue()


def build_client_payment_records_xlsx(*, ctx: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    pink = "F2C4CE"
    soft = "F8FAFC"
    thin = Side(style="thin", color="E2E8F0")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    client = ctx.get("client") or {}
    rows = ctx.get("rows") or []
    totals = ctx.get("totals") or {}

    ws.merge_cells("A1:I1")
    c = ws.cell(row=1, column=1, value=f"{client.get('name') or ''} — Security Officer Payment Records")
    c.font = Font(size=14, bold=True)
    ws.cell(row=2, column=1, value=f"Client Code: {client.get('code') or '—'}").font = Font(color="64748B")

    headers = ["SL", "Security Officer Name", "Address", "Social Security Code",
               "W2", "W9 Direct Deposit", "W9 Zelle Transfer", "W9 Total", "Total (W2+W9)"]
    hr = 4
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=col, value=label)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=pink)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    money_cols = {5, 6, 7, 8, 9}
    r = hr
    for row in rows:
        r += 1
        vals = [row.get("sl"), row.get("officer_name"), row.get("officer_address"),
                row.get("social_security_code"), row.get("w2_amount"),
                row.get("w9_direct_deposit_amount"), row.get("w9_zelle_amount"),
                row.get("w9_total"), row.get("total")]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=(v if v is not None else ""))
            cell.border = border
            if col in money_cols:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right")
    if not rows:
        r += 1
        ws.cell(row=r, column=2, value="No payment records").border = border

    r += 1
    gt = ws.cell(row=r, column=2, value="Grand Total")
    gt.font = Font(bold=True)
    gvals = {5: totals.get("w2", 0), 6: totals.get("w9_direct_deposit", 0),
             7: totals.get("w9_zelle", 0), 8: totals.get("w9_total", 0),
             9: totals.get("grand_total", 0)}
    for col in range(1, 10):
        cell = ws.cell(row=r, column=col, value=gvals.get(col))
        cell.fill = PatternFill("solid", fgColor=soft)
        cell.font = Font(bold=True)
        cell.border = border
        if col in money_cols:
            cell.number_format = '"$"#,##0.00'
            cell.alignment = Alignment(horizontal="right")

    widths = [6, 26, 30, 18, 16, 18, 18, 14, 16]
    for col, w in enumerate(widths, start=1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = ws.cell(row=hr + 1, column=1).coordinate

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()



# ---- Officer detail payment report (dated entries) -------------------------

def build_officer_payment_records_pdf(*, ctx: dict) -> bytes:
    """Per-officer statement.

    Header (centered): client logo, client name, address, phone, email, website.
    Before the table: Officer Name, Social Security Code, Statement Period.
    Table columns: Date, W2, grouped W9 (Direct Deposit / Zelle Transfer /
    W9 Total), Total (W2+W9). Grand totals footer row.
    """
    buf = io.BytesIO()
    half_inch = 0.5 * inch
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=half_inch, rightMargin=half_inch,
        topMargin=half_inch, bottomMargin=half_inch,
    )
    styles = getSampleStyleSheet()
    story = []

    client = ctx.get("client") or {}
    officer = ctx.get("officer") or {}
    records = ctx.get("records") or []
    totals = ctx.get("totals") or {}
    period = ctx.get("period") or {}

    # --- Centered client header ---
    story.extend(_client_center_header(client, doc, styles))
    story.append(Spacer(1, 0.4 * cm))

    # --- Left details ---
    meta_style = ParagraphStyle('meta', parent=styles['Normal'], fontSize=10, alignment=0)
    period_str = f"{period.get('from') or '—'} to {period.get('to') or '—'}"
    meta = [
        [Paragraph("<b>Officer Name:</b>", meta_style), Paragraph(officer.get("name") or "—", meta_style)],
        [Paragraph("<b>Social Security Code:</b>", meta_style), Paragraph(officer.get("social_security_code") or "—", meta_style)],
        [Paragraph("<b>Statement Period:</b>", meta_style), Paragraph(period_str, meta_style)],
    ]
    meta_tbl = Table(meta, colWidths=[4.6 * cm, doc.width - 4.6 * cm], hAlign='LEFT')
    meta_tbl.setStyle(TableStyle([
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3), ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(meta_tbl)
    story.append(Spacer(1, 0.4 * cm))

    # --- Table with grouped W9 header ---
    head = ParagraphStyle('h', parent=styles['Normal'], fontSize=8, leading=10,
                          alignment=1, textColor=colors.HexColor('#0F172A'))
    cell = ParagraphStyle('c', parent=styles['Normal'], fontSize=9, leading=11, alignment=1)

    def P(t):
        return Paragraph(t, cell)

    header_row1 = [Paragraph("<b>Date</b>", head), Paragraph("<b>W2</b>", head),
                   Paragraph("<b>W9</b>", head), "", "", Paragraph("<b>Total (W2+W9)</b>", head)]
    header_row2 = ["", "", Paragraph("<b>Direct Deposit</b>", head),
                   Paragraph("<b>Zelle Transfer</b>", head), Paragraph("<b>W9 Total</b>", head), ""]
    data = [header_row1, header_row2]
    for r in records:
        data.append([
            P(str(r.get("date") or "—")),
            P(f"${r.get('w2_amount', 0):,.2f}"),
            P(f"${r.get('w9_direct_deposit_amount', 0):,.2f}"),
            P(f"${r.get('w9_zelle_amount', 0):,.2f}"),
            P(f"${r.get('w9_total', 0):,.2f}"),
            Paragraph(f"<b>${r.get('total', 0):,.2f}</b>", cell),
        ])
    if not records:
        data.append([P("—"), P("$0.00"), P("$0.00"), P("$0.00"), P("$0.00"), P("$0.00")])
    data.append([
        Paragraph("<b>Grand Total</b>", cell),
        Paragraph(f"<b>${totals.get('w2', 0):,.2f}</b>", cell),
        Paragraph(f"<b>${totals.get('w9_direct_deposit', 0):,.2f}</b>", cell),
        Paragraph(f"<b>${totals.get('w9_zelle', 0):,.2f}</b>", cell),
        Paragraph(f"<b>${totals.get('w9_total', 0):,.2f}</b>", cell),
        Paragraph(f"<b>${totals.get('grand_total', 0):,.2f}</b>", cell),
    ])

    weights = [1.6, 1.3, 1.5, 1.5, 1.3, 1.6]
    total_w = sum(weights)
    col_widths = [doc.width * (w / total_w) for w in weights]
    tbl = Table(data, colWidths=col_widths, repeatRows=2, hAlign='LEFT')
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 1), colors.HexColor('#F2C4CE')),
        ('SPAN', (0, 0), (0, 1)),   # Date
        ('SPAN', (1, 0), (1, 1)),   # W2
        ('SPAN', (2, 0), (4, 0)),   # W9 group
        ('SPAN', (5, 0), (5, 1)),   # Total
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#94A3B8')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F8FAFC')),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.HexColor('#0F172A')),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 0.4 * cm))
    story.append(Paragraph(
        f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        ParagraphStyle('f', parent=styles['Normal'], fontSize=8, textColor=colors.grey)))

    doc.build(story)
    return buf.getvalue()


def build_officer_payment_records_xlsx(*, ctx: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    pink = "F2C4CE"
    soft = "F8FAFC"
    thin = Side(style="thin", color="E2E8F0")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    client = ctx.get("client") or {}
    officer = ctx.get("officer") or {}
    records = ctx.get("records") or []
    totals = ctx.get("totals") or {}
    period = ctx.get("period") or {}

    ws.merge_cells("A1:F1")
    ws.cell(row=1, column=1, value=client.get("name") or "").font = Font(size=14, bold=True)
    ws.cell(row=2, column=1, value="Officer Name").font = Font(bold=True)
    ws.cell(row=2, column=2, value=officer.get("name") or "—")
    ws.cell(row=3, column=1, value="Social Security Code").font = Font(bold=True)
    ws.cell(row=3, column=2, value=officer.get("social_security_code") or "—")
    ws.cell(row=4, column=1, value="Transaction Period").font = Font(bold=True)
    ws.cell(row=4, column=2, value=f"{period.get('from') or '—'} to {period.get('to') or '—'}")

    hr = 6
    # Grouped header: row hr and hr+1
    ws.merge_cells(start_row=hr, start_column=1, end_row=hr + 1, end_column=1)
    ws.merge_cells(start_row=hr, start_column=2, end_row=hr + 1, end_column=2)
    ws.merge_cells(start_row=hr, start_column=3, end_row=hr, end_column=5)
    ws.merge_cells(start_row=hr, start_column=6, end_row=hr + 1, end_column=6)
    ws.cell(row=hr, column=1, value="Date")
    ws.cell(row=hr, column=2, value="W2")
    ws.cell(row=hr, column=3, value="W9")
    ws.cell(row=hr, column=6, value="Total (W2+W9)")
    ws.cell(row=hr + 1, column=3, value="Direct Deposit")
    ws.cell(row=hr + 1, column=4, value="Zelle Transfer")
    ws.cell(row=hr + 1, column=5, value="W9 Total")
    for rr in (hr, hr + 1):
        for col in range(1, 7):
            cell = ws.cell(row=rr, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=pink)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    money_cols = {2, 3, 4, 5, 6}
    r = hr + 1
    for rec in records:
        r += 1
        vals = [rec.get("date"), rec.get("w2_amount"), rec.get("w9_direct_deposit_amount"),
                rec.get("w9_zelle_amount"), rec.get("w9_total"), rec.get("total")]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=(v if v is not None else ""))
            cell.border = border
            if col in money_cols:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="center")
    if not records:
        r += 1
        ws.cell(row=r, column=1, value="—").border = border

    r += 1
    gt = ws.cell(row=r, column=1, value="Grand Total")
    gt.font = Font(bold=True)
    gvals = {2: totals.get("w2", 0), 3: totals.get("w9_direct_deposit", 0),
             4: totals.get("w9_zelle", 0), 5: totals.get("w9_total", 0),
             6: totals.get("grand_total", 0)}
    for col in range(1, 7):
        cell = ws.cell(row=r, column=col, value=gvals.get(col))
        cell.fill = PatternFill("solid", fgColor=soft)
        cell.font = Font(bold=True)
        cell.border = border
        if col in money_cols:
            cell.number_format = '"$"#,##0.00'
            cell.alignment = Alignment(horizontal="right")

    for col, w in enumerate([16, 14, 16, 16, 14, 16], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = ws.cell(row=hr + 2, column=1).coordinate

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

